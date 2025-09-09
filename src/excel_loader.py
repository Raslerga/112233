from __future__ import annotations
from typing import Dict, Optional
from openpyxl import load_workbook


def _normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_inn_to_email_map(excel_path: str) -> Dict[str, str]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    # Пытаемся найти строку заголовков в первых 10 строках
    header_row_idx = None
    for r in range(1, min(10, ws.max_row) + 1):
        row_values = [(_normalize_text(c.value)).lower() for c in ws[r]]
        if any("инн" in v for v in row_values):
            header_row_idx = r
            break

    inn_col_idx: Optional[int] = None
    email_col_idx: Optional[int] = None

    if header_row_idx is not None:
        headers = [(_normalize_text(c.value)).lower() for c in ws[header_row_idx]]
        for idx, name in enumerate(headers, start=1):
            if inn_col_idx is None and ("инн" in name or "inn" in name):
                inn_col_idx = idx
            # Поиск названия колонок для почты
            if email_col_idx is None and any(key in name for key in ("email", "e-mail", "mail", "почт")):
                email_col_idx = idx

    # Фолбэк к B (2) и C (3) по скриншоту, если заголовки не найдены
    if inn_col_idx is None:
        inn_col_idx = 3
    if email_col_idx is None:
        email_col_idx = 2

    mapping: Dict[str, str] = {}

    start_row = (header_row_idx + 1) if header_row_idx else 1
    for r in range(start_row, ws.max_row + 1):
        inn_raw = ws.cell(row=r, column=inn_col_idx).value
        email_raw = ws.cell(row=r, column=email_col_idx).value
        inn = _normalize_text(inn_raw).replace(" ", "")
        email = _normalize_text(email_raw)
        if not inn or not email or "@" not in email:
            continue
        mapping[inn] = email

    return mapping


def find_email_by_inn(excel_path: str, inn: str) -> Optional[str]:
    inn = _normalize_text(inn).replace(" ", "")
    if not inn:
        return None
    mapping = load_inn_to_email_map(excel_path)
    return mapping.get(inn) 